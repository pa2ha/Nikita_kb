from io import BytesIO

from openpyxl import load_workbook

from database.models import Article


async def process_excel_file(file_bytes):
    # Загружаем файл Excel из байтового потока
    workbook = load_workbook(file_bytes)
    
    # Получаем активный лист
    sheet = workbook.active
    
    # Список для хранения объектов модели Article
    articles = []
    
    # Обрабатываем данные из файла Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Предполагаем, что артикул находится в первом столбце (столбец A)
        article = row[0]
        
        # Создаем объект модели Article и добавляем его в список
        new_article = Article(article=article)
        articles.append(new_article)
    
    # Возвращаем список готовых объектов модели Article
    return articles
