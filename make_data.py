import base64
import os
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image


async def update_excel_with_product(filename: str, product_data: dict):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Article", "Created At", "Photo"])

    last_row = ws.max_row + 1
    image_data = base64.b64decode(product_data['photo'])
    img = Image.open(BytesIO(image_data))

    xl_img = XLImage(img)
    ws.column_dimensions['C'].width = 60
    xl_img.anchor = f'C{last_row}'

    
    xl_img.width = 402
    xl_img.height = 409
    
    ws.add_image(xl_img, f"C{last_row}")
    
    ws[f'A{last_row}'] = product_data['article']
    ws[f'B{last_row}'] = product_data['created_at'].strftime("%Y-%m-%d %H:%M:%S")
    
    # Установить ширину столбца для даты
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['A'].width = 25

    # Установить высоту строки для строки с изображением
    ws.row_dimensions[last_row].height = 307.5  # Пример высоты строки
    
    # Сохранить обновленный файл
    wb.save(filename)

